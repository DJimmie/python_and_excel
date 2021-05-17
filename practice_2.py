
from openpyxl import Workbook, load_workbook

excel_file=r'C:\Users\dowdj\OneDrive\Desktop\BBall_Record.xlsx'

wb=load_workbook(excel_file)
ws=wb.active

print(ws['cb16'].value)

ws['cb16'].value=14

wb.save(excel_file)